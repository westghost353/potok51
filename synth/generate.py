"""Генератор синтетических карточек счёта 51 в формате выгрузки 1С 8.3.

Формируется не «чистая таблица», а печатная форма со всеми свойствами
реальной выгрузки: преамбула, двухуровневая шапка с объединёнными ячейками,
разделители по расчётным счетам, многострочные операции, суммы текстом
с неразрывным пробелом, строки итогов.
"""

from __future__ import annotations

import argparse
import calendar
import random
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .profiles import PROFILES, Profile

PERIOD_END = date(2026, 7, 31)
MONTHS = 12

SUPPLIER_NAMES = [
    'ООО «Металлкомплект»', 'ООО «ТД Промресурс»', 'АО «Логистик Экспресс»',
    'ООО «Упаковка Сервис»', 'ООО «Химтрейд»', 'ООО «Автодеталь Плюс»',
    'ООО «Профиль Сталь»', 'ООО «Складские решения»',
]
CUSTOMER_POOL = [
    'ООО «Гранит Строй»', 'ООО «Аквилон Ритейл»', 'АО «Промторг»',
    'ООО «Дельта Маркет»', 'ООО «Сибирь Опт»', 'ООО «Мега Дистрибуция»',
    'ООО «Ремонт и Отделка»', 'ООО «Формат Плюс»', 'АО «Северный Стандарт»',
    'ООО «Крепёж Центр»', 'ООО «Эталон Групп»', 'ООО «Инжиниринг Сервис»',
    'ООО «Азимут Трейд»', 'ООО «Первая Логистическая»', 'ООО «Новый Век»',
    'ООО «Каскад»', 'ООО «Триумф Ритейл»', 'ООО «Базис Комплект»',
    'ООО «Вертикаль»', 'ООО «Стандарт Опт»', 'ООО «Меридиан Трейд»',
    'ООО «Партнёр Групп»', 'ООО «Альфа Комплект»', 'ООО «Резерв Плюс»',
    'ООО «Континент»', 'ООО «Технолайн»', 'ООО «Прогресс Сити»',
    'ООО «Оптима Трейд»', 'ООО «Регион Снаб»', 'ООО «Аметист»',
    'ООО «Бриз Логистик»', 'ООО «Восход Трейд»', 'ООО «Гарант Строй»',
    'ООО «Дом Материалов»', 'ООО «Евро Комплект»',
]
TRANSIT_PARTNERS = [
    'ООО «Стройальянс Капитал»', 'ООО «Промресурс Инвест»', 'ООО «Техноторг Групп»',
    'ООО «Меркурий Логистик»', 'ООО «Аструм Трейд»',
]


@dataclass
class Op:
    dt: datetime
    doc_type: str
    debit_acc: str
    credit_acc: str
    counterparty: str
    contract: str | None
    purpose: str
    amount: float
    account_no: str


def month_starts(end: date, count: int) -> list:
    starts = []
    y, m = end.year, end.month
    for _ in range(count):
        starts.append(date(y, m, 1))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(starts))


def _day(rng: random.Random, month_start: date, lo: int = 1, hi: int = 28) -> datetime:
    last = calendar.monthrange(month_start.year, month_start.month)[1]
    day = min(rng.randint(lo, hi), last)
    return datetime(month_start.year, month_start.month, day,
                    rng.randint(9, 18), rng.randint(0, 59), rng.randint(0, 59))


def build_operations(profile: Profile, seed: int = 20260817) -> tuple[list, list]:
    rng = random.Random(seed)
    starts = month_starts(PERIOD_END, MONTHS)
    acc_main = "40702810" + "".join(str(rng.randint(0, 9)) for _ in range(12))
    acc_second = "40702810" + "".join(str(rng.randint(0, 9)) for _ in range(12))
    accounts = [(acc_main, 'ПАО «Альфа-Банк»')]
    if profile.second_account:
        accounts.append((acc_second, 'ПАО Сбербанк'))

    customers = CUSTOMER_POOL[: profile.customers]
    suppliers = SUPPLIER_NAMES
    ops: list = []

    def add(dt, doc, dacc, cacc, cp, contract, purpose, amount, acc=acc_main):
        if amount <= 0:
            return
        ops.append(Op(dt, doc, dacc, cacc, cp, contract, purpose, round(amount, 2), acc))

    IN = "Поступление на расчетный счет"
    OUT = "Списание с расчетного счета"

    for idx, ms in enumerate(starts):
        season = profile.season[ms.month - 1] if profile.season else 1.0
        revenue = profile.revenue_base * ((1 + profile.revenue_growth) ** idx) * season
        if profile.collapse_factor != 1.0 and idx >= MONTHS - 3:
            revenue *= profile.collapse_factor

        # ---------- выручка ------------------------------------------------
        acq_part = revenue * profile.acquiring_share
        direct_part = revenue - acq_part

        if acq_part > 0:
            settlements = 22
            for k in range(settlements):
                amount = acq_part / settlements * rng.uniform(0.75, 1.3)
                add(_day(rng, ms, 1, 28), IN, "51", "57.03",
                    'АО «Альфа-Банк», эквайринг', None,
                    f"Возмещение по операциям с использованием банковских карт за {ms.strftime('%m.%Y')}",
                    amount)

        if direct_part > 0 and customers:
            top1 = direct_part * profile.top1_share
            rest = direct_part - top1
            add(_day(rng, ms, 3, 25), IN, "51", "62.01", customers[0],
                f"Договор поставки № {100 + idx} от 12.01.2025",
                f"Оплата за товар по счету {rng.randint(100, 999)} от {ms.strftime('%d.%m.%Y')}, в т.ч. НДС 20%",
                top1)
            others = customers[1:] or customers
            weights = [rng.uniform(0.5, 1.5) for _ in others]
            total_w = sum(weights)
            for cp, w in zip(others, weights):
                amount = rest * w / total_w
                if amount < 1000:
                    continue
                add(_day(rng, ms, 1, 28), IN, "51", "62.01", cp,
                    f"Договор № {rng.randint(10, 99)}/{ms.year} от 05.03.2025",
                    f"Оплата по счету {rng.randint(100, 999)} от {ms.strftime('%d.%m.%Y')} за поставленный товар, в т.ч. НДС 20%",
                    amount)

        # ---------- транзит: приток и списание в 1–2 дня --------------------
        transit_volume = 0.0
        if profile.transit_share > 0:
            gross_in = revenue
            transit_volume = gross_in * profile.transit_share / (1 - profile.transit_share)
            n = 14
            for k in range(n):
                amount = transit_volume / n * rng.uniform(0.85, 1.15)
                in_dt = _day(rng, ms, 1, 26)
                add(in_dt, IN, "51", "62.01",
                    CUSTOMER_POOL[rng.randrange(0, 8)],
                    f"Договор № {rng.randint(10, 99)} от 11.02.2026",
                    f"Оплата по договору поставки № {rng.randint(10, 99)}, без НДС",
                    amount)
                out_dt = in_dt + timedelta(days=rng.randint(0, 2), hours=rng.randint(1, 5))
                add(out_dt, OUT, "60.01", "51",
                    TRANSIT_PARTNERS[k % len(TRANSIT_PARTNERS)],
                    f"Договор № {rng.randint(100, 999)} от 20.01.2026",
                    "Оплата по счету за товар, без НДС",
                    amount * rng.uniform(0.990, 0.999))

        # ---------- поставщики ---------------------------------------------
        supplier_total = revenue * profile.supplier_share
        n_sup = 16
        for k in range(n_sup):
            amount = supplier_total / n_sup * rng.uniform(0.6, 1.5)
            add(_day(rng, ms, 2, 27), OUT, "60.01", "51",
                suppliers[k % len(suppliers)],
                f"Договор № {rng.randint(200, 899)} от 14.04.2025",
                f"Оплата по счету {rng.randint(1000, 9999)} за поставленный товар, в т.ч. НДС 20%",
                amount)

        # ---------- фонд оплаты труда и налоги ------------------------------
        payroll = revenue * profile.payroll_share
        if payroll > 0:
            add(datetime(ms.year, ms.month, min(20, calendar.monthrange(ms.year, ms.month)[1]), 11, 0),
                OUT, "70", "51", "Сотрудники организации", None,
                f"Аванс за {ms.strftime('%m.%Y')} по реестру № {idx * 2 + 1}", payroll * 0.4)
            pay_day = min(5, calendar.monthrange(ms.year, ms.month)[1])
            add(datetime(ms.year, ms.month, pay_day, 11, 30),
                OUT, "70", "51", "Сотрудники организации", None,
                f"Заработная плата за {ms.strftime('%m.%Y')} по реестру № {idx * 2 + 2}", payroll * 0.6)
            add(datetime(ms.year, ms.month, min(15, calendar.monthrange(ms.year, ms.month)[1]), 12, 0),
                OUT, "69.01", "51", "УФК по г. Москве (СФР)", None,
                "Страховые взносы по единому тарифу за отчетный период", payroll * 0.30)
            add(datetime(ms.year, ms.month, min(6, calendar.monthrange(ms.year, ms.month)[1]), 12, 15),
                OUT, "68.01", "51", "УФК по г. Москве (ИФНС № 7)", None,
                "НДФЛ с доходов, источником которых является налоговый агент", payroll * 0.13)

        taxes = revenue * profile.tax_share
        add(datetime(ms.year, ms.month, min(28, calendar.monthrange(ms.year, ms.month)[1]), 13, 0),
            OUT, "68.90", "51", "УФК по г. Москве (Казначейство России)", None,
            "Единый налоговый платеж (ЕНС)", taxes)

        # ---------- накладные расходы ---------------------------------------
        if profile.rent_monthly:
            add(_day(rng, ms, 3, 8), OUT, "60.01", "51", 'ООО «Управляющая компания Меркурий»',
                "Договор аренды № 7 от 01.01.2025",
                f"Арендная плата за {ms.strftime('%m.%Y')}, в т.ч. НДС 20%", profile.rent_monthly)
        if profile.comms_monthly:
            add(_day(rng, ms, 5, 12), OUT, "60.01", "51", 'ПАО «Ростелеком»', None,
                f"Услуги связи и интернет за {ms.strftime('%m.%Y')}", profile.comms_monthly)
        if profile.bank_fee_monthly:
            add(_day(rng, ms, 1, 3), OUT, "91.02", "51", 'ПАО «Альфа-Банк»', None,
                "Комиссия за расчетно-кассовое обслуживание", profile.bank_fee_monthly)

        # ---------- наличные и подотчёт --------------------------------------
        if profile.cash_share:
            cash_total = revenue * profile.cash_share
            for k in range(6):
                add(_day(rng, ms, 2, 27), OUT, "71.01", "51",
                    f"Подотчетное лицо {k + 1}", None,
                    "Перечисление в подотчет на хозяйственные нужды", cash_total * 0.55 / 6)
            for k in range(3):
                add(_day(rng, ms, 4, 26), OUT, "50.01", "51", "Касса организации", None,
                    "Снятие наличных по чеку на закупку у населения", cash_total * 0.45 / 3)

        # ---------- собственник ------------------------------------------------
        if profile.owner_share:
            owner_total = revenue * profile.owner_share
            if profile.inn and len(profile.inn) == 12:
                add(_day(rng, ms, 10, 27), OUT, "76.09", "51", profile.org, None,
                    "Перевод собственных средств предпринимателя на личную карту", owner_total)
            else:
                add(_day(rng, ms, 12, 26), OUT, "75.02", "51", "Учредители организации", None,
                    "Выплата дивидендов по решению участника", owner_total)

        # ---------- кредиты ----------------------------------------------------
        if profile.loan_payment:
            add(datetime(ms.year, ms.month, min(25, calendar.monthrange(ms.year, ms.month)[1]), 10, 0),
                OUT, "66.01", "51", 'ПАО «Альфа-Банк»',
                "Кредитный договор № 0012-К от 18.09.2024",
                "Погашение основного долга по кредитному договору", profile.loan_payment)
        if profile.loan_interest:
            add(datetime(ms.year, ms.month, min(25, calendar.monthrange(ms.year, ms.month)[1]), 10, 5),
                OUT, "66.02", "51", 'ПАО «Альфа-Банк»',
                "Кредитный договор № 0012-К от 18.09.2024",
                "Уплата процентов по кредитному договору", profile.loan_interest)

        # ---------- взыскание --------------------------------------------------
        if profile.has_enforcement and idx >= MONTHS - 4:
            add(_day(rng, ms, 14, 22), OUT, "76.02", "51", "Межрайонный ОСП по ЮАО (ФССП)", None,
                "Списание по исполнительному документу № 2-1145/2026", revenue * 0.035)

        # ---------- переводы между своими счетами --------------------------------
        if profile.second_account:
            for _ in range(profile.internal_transfers):
                amount = revenue * 0.08
                dt = _day(rng, ms, 6, 24)
                add(dt, OUT, "51", "51", profile.org, None,
                    "Перевод собственных средств между счетами организации", amount, acc_main)
                add(dt + timedelta(hours=1), IN, "51", "51", profile.org, None,
                    "Пополнение расчетного счета организации", amount, acc_second)

    ops.sort(key=lambda o: (o.account_no, o.dt))
    return ops, accounts


def write_xlsx(profile: Profile, ops: list, accounts: list, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Карточка счета 51"
    bold = Font(bold=True)

    start = ops[0].dt.date().replace(day=1) if ops else date(2025, 8, 1)
    ws.append([profile.org])
    ws.append([f"ИНН {profile.inn}   КПП 771201001"])
    ws.append([f"Карточка счета 51 за {start.strftime('%d.%m.%Y')} - {PERIOD_END.strftime('%d.%m.%Y')}"])
    ws.append([""])          # пустая строка-разделитель (append([]) не создаёт строку)
    header_row = ws.max_row + 1
    ws.append(["Период", "Документ", "Аналитика Дт", "Аналитика Кт",
               "Дебет", "", "Кредит", "", "Текущее сальдо"])
    ws.append(["", "", "", "", "Счет", "Сумма", "Счет", "Сумма", ""])
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=6)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=8)
    for col in (1, 2, 3, 4, 9):
        ws.merge_cells(start_row=header_row, start_column=col,
                       end_row=header_row + 1, end_column=col)
    for row in ws.iter_rows(min_row=header_row, max_row=header_row + 1):
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def money(value: float) -> str:
        whole, frac = f"{abs(value):.2f}".split(".")
        groups = []
        while len(whole) > 3:
            groups.insert(0, whole[-3:])
            whole = whole[:-3]
        groups.insert(0, whole)
        return " ".join(groups) + "," + frac

    balance = profile.opening_balance
    ws.append(["Сальдо на начало", "", "", "", "", money(balance), "", "", money(balance)])

    by_account: dict = {}
    for op in ops:
        by_account.setdefault(op.account_no, []).append(op)

    total_debit = total_credit = 0.0
    for acc, bank in accounts:
        acc_ops = by_account.get(acc, [])
        if not acc_ops:
            continue
        ws.append([f"Расчетный счет: {acc}, {bank}"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for op in acc_ops:
            is_inflow = op.debit_acc == "51" and op.credit_acc != "51"
            if op.debit_acc == "51" and op.credit_acc == "51":
                is_inflow = op.doc_type.startswith("Поступление")
            amount = op.amount
            balance += amount if is_inflow else -amount
            if is_inflow:
                total_debit += amount
            else:
                total_credit += amount

            own = f"Расчетный счет: {acc}"
            cp_lines = [op.counterparty]
            if op.contract:
                cp_lines.append(op.contract)
            cp_lines.append(op.purpose)

            analytics_dt = own if is_inflow else cp_lines[0]
            analytics_kt = cp_lines[0] if is_inflow else own
            doc = f"{op.doc_type} {2000 + len(ws['A']) % 8000:04d}-{op.dt.strftime('%d%m')} от {op.dt.strftime('%d.%m.%Y')}"

            first = [
                op.dt.strftime("%d.%m.%Y %H:%M:%S"), doc, analytics_dt, analytics_kt,
                op.debit_acc, money(amount) if is_inflow else "",
                op.credit_acc, "" if is_inflow else money(amount),
                money(balance),
            ]
            ws.append(first)

            extra = cp_lines[1:]
            if profile.layout == "inline":
                # весь блок аналитики одной ячейкой с переносами строк
                target = 4 if is_inflow else 3
                cell = ws.cell(row=ws.max_row, column=target)
                cell.value = "\n".join(cp_lines)
                cell.alignment = Alignment(wrap_text=True)
            else:
                for line in extra:
                    row = ["", "", "", "", "", "", "", "", ""]
                    row[3 if is_inflow else 2] = line
                    ws.append(row)

        total_debit = round(total_debit, 2)
        total_credit = round(total_credit, 2)

    ws.append(["Обороты за период", "", "", "", "", money(total_debit), "", money(total_credit), ""])
    ws.append(["Сальдо на конец", "", "", "", "", money(balance), "", "", money(balance)])
    for r in (ws.max_row, ws.max_row - 1):
        ws.cell(row=r, column=1).font = bold

    widths = [21, 44, 40, 40, 10, 18, 10, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def generate(profile: Profile, out_dir: Path, seed: int = 20260817) -> Path:
    ops, accounts = build_operations(profile, seed)
    path = out_dir / f"{profile.key}.xlsx"
    write_xlsx(profile, ops, accounts, path)
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Генератор синтетических карточек счёта 51")
    parser.add_argument("--out", default="data/synthetic", help="каталог для файлов")
    parser.add_argument("--seed", type=int, default=20260817)
    parser.add_argument("--profile", help="ключ одного профиля; по умолчанию все")
    args = parser.parse_args()

    out_dir = Path(args.out)
    targets = [p for p in PROFILES if not args.profile or p.key == args.profile]
    for profile in targets:
        path = generate(profile, out_dir, args.seed)
        print(f"{profile.key:26s} -> {path}")


if __name__ == "__main__":
    main()
