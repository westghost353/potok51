"""Выгрузка результата анализа в Excel: шесть листов для ручной проверки."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import INDUSTRY_RU
from ..models import CATEGORY_RU, DECISION_RU, Analysis

FILLS = {
    "GREEN": PatternFill("solid", fgColor="D6F0DF"),
    "AMBER": PatternFill("solid", fgColor="FBECC8"),
    "RED": PatternFill("solid", fgColor="F7D4CE"),
    "NA": PatternFill("solid", fgColor="EDEFF2"),
}


def _autosize(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header(ws, values: list) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_excel(a: Analysis, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Решение"
    ws.append(["Поток 51 — результат кредитного анализа"])
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Клиент", a.meta.organization or "—"),
        ("ИНН", a.meta.inn or "—"),
        ("Отрасль", INDUSTRY_RU.get(a.industry, a.industry)),
        ("Период", f"{a.monthly[0].month}—{a.monthly[-1].month}" if a.monthly else "—"),
        ("Решение", DECISION_RU[a.decision.code.value]),
        ("Лимит, ₽", a.limit.final),
        ("Диапазон, ₽", f"{a.limit.range_low:,.0f} — {a.limit.range_high:,.0f}".replace(",", " ")),
        ("Связывающий ограничитель", a.limit.binding_constraint or "—"),
        ("Множитель флагов", a.limit.multiplier),
        ("Очищенная выручка за период, ₽", a.volumes.adjusted_revenue),
        ("Свободный операционный поток, ₽", a.volumes.fcf),
        ("Доля автоклассификации", a.metrics.get("auto_classified_share")),
        ("Версия правил", a.rules_version),
        ("SHA-256 файла", a.source_sha256),
    ]
    for label, value in rows:
        ws.append([label, value])
    for factor in a.decision.stop_factors:
        ws.append(["Стоп-фактор", factor])
    for reason in a.decision.reasons:
        ws.append(["Основание", reason])
    _autosize(ws, [38, 76])

    ws = wb.create_sheet("Индикаторы")
    _header(ws, ["Код", "Индикатор", "Значение", "Статус", "Пояснение"])
    for i in a.indicators:
        ws.append([i.code, i.name, i.display, i.status.value, i.explanation])
        ws.cell(row=ws.max_row, column=4).fill = FILLS.get(i.status.value, FILLS["NA"])
    _autosize(ws, [8, 34, 16, 12, 90])

    ws = wb.create_sheet("Помесячно")
    _header(ws, ["Месяц", "Валовой приток", "Квалифицированный приток",
                 "Очищенная выручка", "Операционный отток", "Свободный поток"])
    for m in a.monthly:
        ws.append([m.month, m.gross_inflow, m.qualified_inflow,
                   m.adjusted_revenue, m.operating_outflow, m.fcf])
    _autosize(ws, [12, 20, 26, 22, 22, 20])

    ws = wb.create_sheet("Исключения")
    _header(ws, ["Категория", "Что исключено", "Сумма, ₽", "Доля притока", "Операций"])
    for e in a.volumes.exclusions:
        ws.append([e.category, e.label, e.amount, e.share, len(e.rows)])
    _autosize(ws, [26, 36, 18, 14, 12])

    ws = wb.create_sheet("Операции")
    _header(ws, ["Строка", "Дата", "Счёт", "Корсчёт", "Контрагент", "Договор",
                 "Назначение", "Приход", "Расход", "Категория", "Источник разметки",
                 "Исключено как", "Пара"])
    for t in a.transactions:
        ws.append([
            t.row_no, t.date, t.account_no, t.corr_account, t.counterparty, t.contract,
            t.purpose, t.inflow or None, t.outflow or None,
            CATEGORY_RU.get(t.category.value, t.category.value),
            t.category_source, t.excluded_reason, t.link_id,
        ])
    ws.freeze_panes = "A2"
    _autosize(ws, [8, 12, 22, 10, 32, 26, 60, 16, 16, 26, 26, 18, 12])

    ws = wb.create_sheet("Неклассифицированное")
    _header(ws, ["Строка", "Дата", "Направление", "Контрагент", "Назначение", "Корсчёт", "Сумма, ₽"])
    for u in a.unclassified_top:
        ws.append([u["row_no"], u["date"], u["direction"], u["counterparty"],
                   u["purpose"], u["corr_account"], u["amount"]])
    _autosize(ws, [8, 12, 14, 32, 60, 10, 16])

    ws = wb.create_sheet("Качество данных")
    _header(ws, ["Код", "Проверка", "Критическая", "Пройдена", "Детали"])
    for c in a.data_quality.checks:
        ws.append([c.code, c.name, "да" if c.critical else "нет",
                   "да" if c.passed else "нет", c.detail])
    _autosize(ws, [8, 34, 14, 12, 90])

    wb.save(path)
    return path
