"""Чтение файла в плоскую сетку строк с развёрнутыми объединёнными ячейками.

Объединённые ячейки в выгрузках 1С — норма, а не исключение: openpyxl
возвращает значение только в левой верхней ячейке диапазона, остальные — None.
Здесь значение размножается на весь диапазон, иначе колонки «съезжают».
"""

from __future__ import annotations

from pathlib import Path

from .base import ReadError


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\xa0", " ").strip()
    return str(value).strip()


def load_grid(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx(path)
    if suffix == ".xls":
        return _load_xls(path)
    raise ReadError(f"Неподдерживаемое расширение файла: {suffix}")


def _load_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    width = ws.max_column or 1
    grid = [
        [_cell_text(c.value) for c in row]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=width)
    ]
    for rng in ws.merged_cells.ranges:
        top_left = grid[rng.min_row - 1][rng.min_col - 1]
        if not top_left:
            continue
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                if r < len(grid) and c < len(grid[r]):
                    grid[r][c] = top_left
    wb.close()
    return grid


def _load_xls(path: Path) -> list[list[str]]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheet = book.sheet_by_index(0)
    grid = [
        [_cell_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]
    for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []):
        top_left = grid[rlo][clo]
        if not top_left:
            continue
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                grid[r][c] = top_left
    return grid
